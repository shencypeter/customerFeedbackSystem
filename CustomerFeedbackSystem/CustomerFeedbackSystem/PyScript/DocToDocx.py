# 安裝套件：pip install openpyxl pywin32
# 用法：python D:\系統資料\project\itriDoc\CustomerFeedbackSystem\CustomerFeedbackSystem\PyScript\DocToDocx.py "C:\Users\peter\OneDrive\Desktop\待轉換清單.xlsx" -o "C:\Users\peter\OneDrive\Desktop\GMP文件庫(轉換doc後)" -s 1 -c A


# 用法：python D:\系統資料\project\itriDoc\CustomerFeedbackSystem\CustomerFeedbackSystem\PyScript\DocToDocx.py "C:\Users\peter\OneDrive\Desktop\重新轉docx.xlsx" -o "C:\Users\peter\OneDrive\Desktop\GMP文件庫(轉換doc後)" -s 1 -c A


# -*- coding: utf-8 -*-
import argparse
import os
from pathlib import Path
from typing import Iterable, Union, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 只有 Windows + 已安裝 Word 才可用
try:
    import win32com.client as win32
    from win32com.client import constants
except Exception as e:
    # 延後到執行轉檔時再提示也行，但這裡先提醒
    pass


def read_paths_from_excel(
    excel_path: Path,
    sheet: Union[int, str] = 1,
    col_letter: str = "A",
    skip_header: bool = False,  # 可選：若第1列是標題可略過
):
    """
    從 Excel 指定工作表與欄位抓出每一列的完整檔案路徑字串，轉為 Path。
    空白列、None、自動略過；去除前後空白與引號。
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet - 1] if isinstance(sheet, int) else wb[sheet]

    col_letter = col_letter.upper().strip()
    col_idx = column_index_from_string(col_letter)

    for idx, row in enumerate(
        ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True), start=1
    ):
        if skip_header and idx == 1:
            continue
        val = row[0]
        if val is None:
            continue
        s = str(val).strip().strip('"').strip("'")
        if not s:
            continue
        yield Path(s)



def unique_target_path(dst_dir: Path, base_name: str, ext: str) -> Path:
    """
    產生不重覆的輸出路徑：若 foo.docx 已存在，就變成 foo (2).docx、foo (3).docx ...
    """
    p = dst_dir / f"{base_name}{ext}"
    if not p.exists():
        return p
    i = 2
    while True:
        q = dst_dir / f"{base_name} ({i}){ext}"
        if not q.exists():
            return q
        i += 1


def convert_doc_to_docx(
    word_app,
    src_path: Path,
    dst_path: Path,
    macro_enabled: bool = False,
) -> Tuple[bool, str]:
    """
    使用 Word COM 將 .doc 轉 .docx（或 .docm），回傳 (成功與否, 訊息)。
    """
    try:
        # 常數：docx=12, docm=13
        file_format = 13 if macro_enabled else 12

        doc = word_app.Documents.Open(str(src_path))
        # SaveAs2 可避免舊版相容問題
        doc.SaveAs2(str(dst_path), FileFormat=file_format)
        doc.Close(SaveChanges=False)
        return True, "OK"
    except Exception as e:
        try:
            # 若文件有開啟，確保關閉
            doc.Close(SaveChanges=False)  # type: ignore
        except Exception:
            pass
        return False, f"ERROR: {e}"


def main():
    parser = argparse.ArgumentParser(
        description="根據 Excel 內的檔案完整路徑清單，將 .doc 轉為 .docx 並另存在新資料夾。"
    )
    parser.add_argument("excel", help="Excel 檔（.xlsx/.xlsm），A 欄放完整檔案路徑")
    parser.add_argument(
        "-o", "--outdir", required=True, help="輸出資料夾（會自動建立）"
    )
    parser.add_argument(
        "-s",
        "--sheet",
        help="工作表索引(從1起算)或工作表名稱，預設 1",
        default="1",
    )
    parser.add_argument(
        "-c", "--column", help="欄位字母（預設 A）", default="A"
    )
    parser.add_argument(
        "--copy-docx",
        action="store_true",
        help="若來源已是 .docx，是否直接複製到輸出資料夾（預設略過）",
    )
    parser.add_argument(
        "--macro-enabled",
        action="store_true",
        help="輸出為 .docm（適用含巨集檔案），預設為 .docx",
    )
    
    parser.add_argument("--skip-header", action="store_true", help="A 欄首列是標題時啟用")
    
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        raise SystemExit(f"找不到 Excel 檔：{excel_path}")

    out_dir = Path(args.outdir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # 解析 sheet 參數
    sheet_arg: Union[int, str]
    if args.sheet.isdigit():
        sheet_arg = int(args.sheet)
    else:
        sheet_arg = args.sheet

    # 讀取路徑清單
    paths = list(read_paths_from_excel(excel_path, sheet=sheet_arg, col_letter=args.column, skip_header=args.skip_header))
    if not paths:
        raise SystemExit("Excel 指定欄位沒有可用的路徑資料。")

    # 啟動 Word（隱藏、關閉警告）
    try:
        word = win32.Dispatch("Word.Application")
    except Exception as e:
        raise SystemExit(
            "無法啟動 Microsoft Word。請確認本機已安裝 Office/Word，且 Python/Office 位元數相容（建議皆為 64 位）。"
        ) from e

    word.Visible = False
    # 0 = wdAlertsNone
    try:
        word.DisplayAlerts = 0
    except Exception:
        pass

    total = len(paths)
    converted = 0
    skipped = 0
    failed = 0

    for i, src in enumerate(paths, start=1):
        print(f"[{i}/{total}] 處理：{src}")
        if not src.exists():
            print("  -> 路徑不存在，略過。")
            failed += 1
            continue

        ext = src.suffix.lower()
        if ext == ".doc":
            # 目標路徑：同名但副檔名 .docx 或 .docm
            base = src.stem
            target_ext = ".docm" if args.macro_enabled else ".docx"
            dst = unique_target_path(out_dir, base, target_ext)

            ok, msg = convert_doc_to_docx(
                word, src, dst, macro_enabled=args.macro_enabled
            )
            if ok:
                converted += 1
                print(f"  -> 轉檔成功：{dst}")
            else:
                failed += 1
                print(f"  -> 轉檔失敗：{msg}")

        elif ext == ".docx":
            if args.copy_docx:
                dst = unique_target_path(out_dir, src.stem, ".docx")
                try:
                    # 複製已是 docx 的檔案（可選）
                    from shutil import copy2
                    copy2(src, dst)
                    skipped += 1
                    print(f"  -> 已是 .docx，複製到：{dst}")
                except Exception as e:
                    failed += 1
                    print(f"  -> 複製失敗：{e}")
            else:
                skipped += 1
                print("  -> 已是 .docx，略過。")
        else:
            skipped += 1
            print("  -> 非 .doc/.docx，略過。")

    # 關閉 Word
    try:
        word.Quit()
    except Exception:
        pass

    print("\n==== 結果 ====")
    print(f"總計：{total}")
    print(f"成功轉檔：{converted}")
    print(f"略過：{skipped}")
    print(f"失敗：{failed}")
    print(f"輸出資料夾：{out_dir}")


if __name__ == "__main__":
    main()
