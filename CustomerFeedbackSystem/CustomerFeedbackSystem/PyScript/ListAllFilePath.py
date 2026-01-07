# 安裝套件：pip install openpyxl
# 用法：python D:\系統資料\project\itriDoc\CustomerFeedbackSystem\CustomerFeedbackSystem\PyScript\ListAllFilePath.py "C:\Users\peter\OneDrive\Desktop\GMP文件庫(2020NEW)" -o "C:\Users\peter\OneDrive\Desktop\原始輸出清單.xlsx"

# 用法：python D:\系統資料\project\itriDoc\CustomerFeedbackSystem\CustomerFeedbackSystem\PyScript\ListAllFilePath.py "C:\Users\peter\OneDrive\Desktop\GMP文件庫(合併與重新命名)" -o "C:\Users\peter\OneDrive\Desktop\合併後輸出清單.xlsx"

# -*- coding: utf-8 -*-
import argparse
import os
from pathlib import Path
from datetime import datetime
from typing import Iterator, Tuple, List, Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit("請先安裝 openpyxl： pip install openpyxl") from e


def iter_entries(root: Path) -> Iterator[Tuple[Path, os.DirEntry, bool]]:
    """
    以 os.scandir 遞迴掃描，比 os.walk 在大量檔案時更快。
    產出 (current_dir, entry, is_dir)。
    """
    stack = [root]
    while stack:
        current = stack.pop()
        try:
            with os.scandir(current) as it:
                for entry in it:
                    try:
                        is_dir = entry.is_dir(follow_symlinks=False)
                    except OSError:
                        # 權限不足/破損捷徑
                        continue
                    yield current, entry, is_dir
                    if is_dir:
                        stack.append(Path(entry.path))
        except (PermissionError, FileNotFoundError):
            # 略過無權限或瞬間被移除的目錄
            continue


def safe_stat(entry: os.DirEntry):
    try:
        return entry.stat(follow_symlinks=False)
    except (PermissionError, FileNotFoundError, OSError):
        return None


def excel_autofit(ws):
    # 粗略自動欄寬：取每欄字元長度上限（含標題）
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = cell.value
            if val is None:
                continue
            text = str(val)
            if len(text) > max_len:
                max_len = len(text)
        # 預留一些空間
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 80)


def export_to_excel(rows: List[List[Any]], headers: List[str], out_path: Path):
    wb: Workbook = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    ws.append(headers)
    for r in rows:
        ws.append(r)

    # 凍結首列 & 篩選
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    excel_autofit(ws)

    # 將日期欄位套格式
    date_cols = {"ModifiedTime", "CreatedTime"}
    for col_idx, header in enumerate(headers, start=1):
        if header in date_cols:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(
        description="掃描目標資料夾並匯出所有資料夾/檔案清單為 Excel"
    )
    parser.add_argument("target", type=str, help="要掃描的資料夾路徑")
    parser.add_argument(
        "-o", "--output", type=str, default=None, help="輸出 Excel 檔（.xlsx）路徑"
    )
    args = parser.parse_args()

    root = Path(args.target).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        raise SystemExit(f"找不到資料夾：{root}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else Path.cwd() / f"folder_inventory_{timestamp}.xlsx"
    )
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")

    headers = [
        "Name",
        "Type",
        "Extension",
        "Size(bytes)",
        "ModifiedTime",
        "CreatedTime",
        "RelativePath",
        "Parent",
        "Depth",
        "FullPath",
    ]

    rows: List[List[Any]] = []
    root_len = len(str(root))

    for current, entry, is_dir in iter_entries(root):
        st = safe_stat(entry)
        name = entry.name

        full_path = Path(entry.path)
        try:
            rel = full_path.relative_to(root)
        except ValueError:
            # 理論上不會發生，但以防萬一
            rel = full_path

        depth = len(rel.parts)
        parent = rel.parent.as_posix() if rel.parent != Path(".") else "/"

        # 基本欄位
        typ = "Folder" if is_dir else "File"
        ext = ("" if is_dir else Path(name).suffix).lower()
        size = (None if is_dir else (st.st_size if st else None))

        mtime = datetime.fromtimestamp(st.st_mtime) if (st and st.st_mtime) else None
        ctime = datetime.fromtimestamp(st.st_ctime) if (st and st.st_ctime) else None

        rows.append(
            [
                name,
                typ,
                ext,
                size,
                mtime,
                ctime,
                rel.as_posix(),
                parent,
                depth,
                str(full_path),
            ]
        )

    # Excel 行數上限提示
    if len(rows) + 1 > 1_048_576:
        print("警告：資料列超過 Excel 上限，部分資料可能無法完整寫入。請改用 CSV 匯出流程。")

    export_to_excel(rows, headers, out_path)
    print(f"已完成，輸出檔：{out_path}")


if __name__ == "__main__":
    main()
