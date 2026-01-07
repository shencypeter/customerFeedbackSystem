# 安裝套件：pip install openpyxl
# 用法：python D:\系統資料\project\itriDoc\BioMedDocManager\BioMedDocManager\PyScript\CopyOthers.py "C:\Users\peter\OneDrive\Desktop\剩下的.xlsx" -o "C:\Users\peter\OneDrive\Desktop\GMP文件庫(剩下的)" -s 1 -c A

# -*- coding: utf-8 -*-
import argparse
from pathlib import Path
from typing import Iterable, Union
from shutil import copy2
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def read_paths_from_excel(
    excel_path: Path,
    sheet: Union[int, str] = 1,
    col_letter: str = "A",
    skip_header: bool = False,
) -> Iterable[Path]:
    """
    從 Excel 指定工作表與欄位抓出每列完整檔案路徑。
    會自動去除前後空白與引號；空白列略過。
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet - 1] if isinstance(sheet, int) else wb[sheet]

    col_idx = column_index_from_string(col_letter.upper().strip())
    for idx, row in enumerate(ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True), start=1):
        if skip_header and idx == 1:
            continue
        val = row[0]
        if val is None:
            continue
        s = str(val).strip().strip('"').strip("'")
        if not s:
            continue
        yield Path(s)

def unique_target_path(dst_dir: Path, base_name: str, ext: str) -> Path:
    """
    產生不重覆的輸出路徑：foo.txt -> foo (2).txt -> foo (3).txt ...
    """
    p = dst_dir / f"{base_name}{ext}"
    if not p.exists():
        return p
    i = 2
    while True:
        q = dst_dir / f"{base_name} ({i}){ext}"
        if not q.exists():
            return q
        i += 1

def main():
    parser = argparse.ArgumentParser(description="讀 Excel 路徑清單，將檔案複製到指定資料夾。")
    parser.add_argument("excel", help="Excel 檔（.xlsx/.xlsm），指定欄位放完整檔案路徑")
    parser.add_argument("-o", "--outdir", required=True, help="輸出資料夾（會自動建立）")
    parser.add_argument("-s", "--sheet", default="1", help="工作表索引(從1起算) 或 名稱，預設 1")
    parser.add_argument("-c", "--column", default="A", help="路徑所在欄位字母，預設 A")
    parser.add_argument("--skip-header", action="store_true", help="若第1列為表頭請加入此旗標")
    parser.add_argument("--overwrite", action="store_true", help="若輸出已存在同名檔，直接覆蓋（預設改名避免覆蓋）")
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        raise SystemExit(f"找不到 Excel 檔：{excel_path}")

    out_dir = Path(args.outdir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # 解析 sheet 參數（數字或名稱）
    sheet_arg: Union[int, str] = int(args.sheet) if args.sheet.isdigit() else args.sheet

    paths = list(read_paths_from_excel(excel_path, sheet=sheet_arg, col_letter=args.column, skip_header=args.skip_header))
    if not paths:
        raise SystemExit("Excel 指定欄位沒有可用的路徑資料。")

    total = len(paths)
    copied = 0
    skipped = 0
    failed = 0

    for i, src in enumerate(paths, start=1):
        print(f"[{i}/{total}] 處理：{src}")
        try:
            if not src.exists():
                print("  -> 路徑不存在，略過。")
                failed += 1
                continue
            if not src.is_file():
                print("  -> 不是檔案（可能是資料夾），略過。")
                skipped += 1
                continue

            base = src.stem
            ext = src.suffix
            if args.overwrite:
                dst = out_dir / f"{base}{ext}"
            else:
                dst = unique_target_path(out_dir, base, ext)

            copy2(src, dst)  # copy2 會連同時間戳等中繼資料
            print(f"  -> 已複製到：{dst}")
            copied += 1

        except Exception as e:
            print(f"  -> 複製失敗：{e}")
            failed += 1

    print("\n==== 結果 ====")
    print(f"總計：{total}")
    print(f"成功複製：{copied}")
    print(f"略過：{skipped}")
    print(f"失敗：{failed}")
    print(f"輸出資料夾：{out_dir}")

if __name__ == "__main__":
    main()
